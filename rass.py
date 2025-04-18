# rass.py

import os

import smtplib
import imaplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import Header
import openpyxl
import xlrd



def read_value_from_txt(txtname):
    values = {}
    with open(txtname, 'r', encoding='utf-8') as file:
        current_key = None
        current_value = []

        for line in file:
            line = line.strip()
            line = line.split('#')[0].strip()

            if not line:
                continue

            if ':' in line:
                if current_key is not None:
                    values[current_key] = ' '.join(current_value).strip().rstrip("'")  # Убираем лишнюю кавычку

                key, value = line.split(":", 1)
                current_key = key.strip()
                current_value = [value.strip().strip("'")]

            else:
                current_value.append(line.strip())

        if current_key is not None:
            values[current_key] = ' '.join(current_value).strip().rstrip("'")

        for key, value in values.items():
            if value.isdigit():
                values[key] = int(value)
            else:
                try:
                    values[key] = float(value)
                except ValueError:
                    pass

            if value.startswith("http://") or value.startswith("https://"):
                pass
            elif "%aw_random%" in value:
                pass

    return values


def values_from_txt(txtname):
    txtname = 'config.txt'
    values = read_value_from_txt(txtname)
    user = values['user']
    app_password = values['app_password']
    text = values['body']
    signature = values['signature']
    topic = values['topic']
    pixel = values['pixel']
    return sender_to,name_list,user,app_password,text,signature,topic,pixel

file_count = 0
sender_to = []
name_list = []

################################################    select_db_file

# Выбор файла(не более 1) и выбор типа обработки(в зависимости от формата)
def excel_format(db_file_names,db_filenames):
    if len(db_file_names) == 1:
        for file in db_file_names:
            file = file.split('.')[1].strip()
            if file == "xlsx":
                sender_to.clear(),name_list.clear()
                workbook = openpyxl.load_workbook(db_filenames)
                active_workbook = workbook.active
                email_info_xlsx(active_workbook)
                print(f"Выбран файл{db_file_names[0]}")
                print(name_list,sender_to)
                return workbook.active,True
            if file == "xls":
                sender_to.clear(),name_list.clear()
                workbook = xlrd.open_workbook(db_filenames)
                sheet = workbook.sheet_by_index(0)
                email_info_xls(sheet)
                print(f"Выбран файл{db_file_names[0]}")
                print(name_list, sender_to)
                return sheet,True
    else:
        print("Выберано 2 и более файлов")
        return False

# Обработка XlSX
def email_info_xlsx(active_workbook):
    for i in range(1, active_workbook.max_row + 1):
        sender_to.append(active_workbook.cell(i, 1).value)
        name_list.append(active_workbook.cell(i, 2).value)
    return sender_to, name_list

# Обработка XLS
def email_info_xls(sheet):
    for i in range(sheet.nrows):
        sender_to.append(sheet.cell_value(i, 0))
        name_list.append(sheet.cell_value(i, 1))
    return sender_to, name_list

################################################    select_db_file


def attach_file_save(filenames):
    global file_count
    file_count = len(filenames)
    filename_t = filenames
    print(filename_t)
    return filename_t


def sendd(update_progress, filename_t, thread):
    global file_count

    sender_to, name_list, user, app_password, text, signature, topic, pixel = values_from_txt("config.txt")

    # Настройка SMTP и IMAP
    domain = user.split('@')[-1]
    smtp_imap_map = {
        'gmail.com': ('smtp.gmail.com', 'imap.gmail.com'),
        'mail.ru':   ('smtp.mail.ru', 'imap.mail.ru'),
        'bk.ru':     ('smtp.mail.ru', 'imap.mail.ru'),
        'yandex.ru': ('smtp.yandex.ru', 'imap.yandex.ru')
    }
    mail, imamail = smtp_imap_map.get(domain, (None, None))

    if not mail or not imamail:
        print(f"Неизвестный домен: {domain}")
        return

    try:
        smtp_server = smtplib.SMTP_SSL(mail, 465)
        smtp_server.login(user, app_password)

        imap_server = imaplib.IMAP4_SSL(imamail)
        imap_server.login(user, app_password)

        status, folders = imap_server.list()
        sent_folder = "Sent"
        for folder in folders:
            decoded = folder.decode()
            if '\\Sent' in decoded:
                sent_folder = decoded.split(' "/" ')[-1].strip('"')
                break

        email_itera = int(100 / len(sender_to))
        email_number = 0
        email_maximum = len(sender_to)
        maximum = email_itera * email_maximum

        for i, name in zip(sender_to, name_list):
            if not thread.is_running:
                break

            name = name if name else " "

            # --- Формирование письма ---
            msg = MIMEMultipart()
            msg['From'] = user
            msg['To'] = i
            msg['Subject'] = topic

            body = text.replace('{name_list}', name)
            pixel_watch = f"""
                <img src="https://script.google.com/macros/s/AKfycbyt8txkXh775E0sofXYbnzlq11PTkSa2Cfvy-To2XcHSB46Iv64T6Elvx8ob3wX8lYo/exec?user_id={i}" 
                     alt="" 
                     style="border: none; display: none;" />
            """

            msg.attach(MIMEText(body, 'html', 'utf-8'))
            msg.attach(MIMEText(signature, 'html', 'utf-8'))
            msg.attach(MIMEText(pixel_watch, 'html', 'utf-8'))

            # Вложения
            for attach in filename_t:
                with open(attach, 'rb') as f:
                    part = MIMEApplication(f.read(), Name=os.path.basename(attach))
                    part['Content-Disposition'] = f'attachment; filename="{Header(os.path.basename(attach), "utf-8").encode()}"'
                    msg.attach(part)

            # Отправка
            try:
                smtp_server.sendmail(user, i, msg.as_string())

                # Найти имя папки с тегом \Sent
                status, folders = imap_server.list()
                sent_folder = None
                for folder in folders:
                    decoded = folder.decode()
                    if '\\Sent' in decoded:
                        sent_folder = decoded.split(' "/" ')[-1].strip('"')
                        break

                # Безопасное добавление письма в отправленные
                if sent_folder:
                    if 'gmail.com' in user:
                        pass
                    else:
                        imap_server.append(sent_folder, [], None, msg.as_bytes())
                else:
                    print("Не удалось найти папку Sent — письмо не добавлено в отправленные")
                print(f"Письмо отправлено: {i}")
            except Exception as e:
                print(f"Ошибка при отправке на {i}: {e} {sent_folder}")

            email_number += 1
            email_progress_value = email_itera * email_number
            update_progress(email_progress_value, maximum, email_maximum, email_number)

        smtp_server.quit()
        imap_server.logout()

    except Exception as e:
        print(f"Ошибка подключения: {e}")

